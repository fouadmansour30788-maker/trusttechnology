"""Parse FULL PRICE LIST 11062026.xlsx into the site's product catalog.

Emits:
  - src/data/products.ts          (typed catalog for immediate use)
  - supabase/migrations/002_seed_products.sql

Run from the project root:  python scripts/parse-pricelist.py "<path-to-xlsx>"
"""
import sys, re, json, io, os
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else \
    os.path.expanduser("~/OneDrive/Desktop/FULL PRICE LIST 11062026.xlsx")
HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ── helpers ──────────────────────────────────────────────────────────────
def slugify(s):
    s = re.sub(r"[^a-z0-9]+", "-", str(s).lower()).strip("-")
    return s[:60] or "item"

def clean(v):
    if v is None:
        return None
    s = str(v).replace("\n", " ").replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s or None

def parse_price(v):
    """Return (price_float_or_0, on_request_bool)."""
    if v is None:
        return 0.0, True
    s = str(v).strip().upper()
    if "CALL" in s or s in ("", "-", "TBA"):
        return 0.0, True
    m = re.search(r"\d+(?:\.\d+)?", str(v).replace(",", ""))
    return (float(m.group()), False) if m else (0.0, True)

def is_section_header(a, b):
    if not a:
        return False
    a = str(a)
    return (b is None and not any(c.islower() for c in a)
            and len(a) < 28 and "PRICE" not in a and "Availab" not in a
            and "VAT" not in a)

BRAND_NORM = {
    "MACBOOK NEO": "Apple", "MACBOOK AIR": "Apple", "MACBOOK PRO": "Apple",
    "LENOVO DESKTOP": "Lenovo", "NEC": "NEC", "ACER": "Acer", "DELL": "Dell",
    "HP": "HP", "LENOVO": "Lenovo", "MSI": "MSI", "GIGABYTE": "Gigabyte",
}
SECTIONS = {"DESKTOP", "SERVER"}

# ── parse ────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX, data_only=True)
def rows(ws):
    return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]
sheets = {ws.title: rows(ws) for ws in wb.worksheets}

products = []
seen_slugs = {}

HEADER_WORDS = {"MODEL", "MODEL NUMBER", "ITEM", "PICTURE", "POS", "CPU",
                "DESCRIPTION", "PRICE", "PRICE HT"}

def add(name, category, brand, ptype, specs, price_val, sku=None):
    name = clean(name)
    if not name or len(name) < 3 or name.upper() in HEADER_WORDS:
        return
    price, on_request = parse_price(price_val)
    base = slugify((brand + "-" if brand else "") + name)
    slug = base
    seen_slugs[base] = seen_slugs.get(base, 0) + 1
    if seen_slugs[base] > 1:
        slug = f"{base}-{seen_slugs[base]}"
    specs = {k: clean(v) for k, v in specs.items() if clean(v)}
    tags = []
    if brand:
        tags.append({"name": brand, "slug": slugify(brand), "type": "brand"})
    if ptype:
        tags.append({"name": ptype.title(), "slug": slugify(ptype), "type": "type"})
    products.append({
        "name": name, "slug": slug, "category": category, "brand": brand,
        "price": price, "priceOnRequest": on_request,
        "sku": clean(sku), "specs": specs, "tags": tags,
    })

# LAPTOPS -----------------------------------------------------------------
LAP = {1: "CPU", 2: "RAM", 3: "Storage", 4: "VGA", 5: "Screen",
       6: "LAN", 7: "Keyboard", 8: "OS", 9: "Warranty"}
brand = None
for r in sheets["LAPTOPS"]:
    a = clean(r[0])
    if is_section_header(r[0], r[1] if len(r) > 1 else None) and a in BRAND_NORM:
        brand = BRAND_NORM[a]; continue
    if a and (len(r) > 1 and r[1]):  # has CPU → product row
        specs = {LAP[i]: r[i] for i in LAP if i < len(r)}
        add(a, "laptops", brand, "laptop", specs, r[10] if len(r) > 10 else None)

# DESKTOPS-SERVER ---------------------------------------------------------
DESK = {1: "CPU", 2: "RAM", 3: "Storage", 4: "Keyboard/Mouse",
        5: "Graphics", 8: "OS", 9: "Warranty"}
brand, section = None, "DESKTOP"
for r in sheets["DESKTOPS-SERVER"]:
    a = clean(r[0])
    if is_section_header(r[0], r[1] if len(r) > 1 else None):
        if a in SECTIONS:
            section = a
        elif a in BRAND_NORM:
            brand = BRAND_NORM[a]
        continue
    if a and (len(r) > 1 and r[1]):
        cat = "desktops" if section == "DESKTOP" else "professional"
        specs = {DESK[i]: r[i] for i in DESK if i < len(r)}
        add(a, cat, brand, "desktop", specs, r[10] if len(r) > 10 else None)

# PHILIPS (monitors) ------------------------------------------------------
for r in sheets["PHILIPS"]:
    model = clean(r[1]) if len(r) > 1 else None
    if not model or "WARRANTY" in str(model).upper():
        continue
    specs = {"Description": r[2] if len(r) > 2 else None,
             "Connectivity": r[3] if len(r) > 3 else None}
    add(model, "monitors", "Philips", "monitor", specs,
        r[4] if len(r) > 4 else None, sku=model)

# COMPATIBLE TONER --------------------------------------------------------
for r in sheets["COMPATIBLE TONER"]:
    a = clean(r[0])
    if not a or a.upper() in ("ITEM",) or "Availab" in str(a):
        continue
    specs = {"Capacity": r[1] if len(r) > 1 else None,
             "Pcs/Carton": r[2] if len(r) > 2 else None}
    add(a, "ink-toner", "Cedar", "toner", specs, r[3] if len(r) > 3 else None)

# POS-BARCODE -------------------------------------------------------------
section = None
for r in sheets["POS-BARCODE"]:
    a = clean(r[0])
    b = clean(r[1]) if len(r) > 1 else None
    if a and b in ("CASH HT", "PRICE", "HT"):
        section = a; continue
    if a and ("Availab" in a or "VAT" in a.upper()):
        continue
    if a and b:
        add(a, "pos-systems", None, "pos", {"Category": section}, b)

# ── stats ────────────────────────────────────────────────────────────────
from collections import Counter
cat_counts = Counter(p["category"] for p in products)
call_count = sum(1 for p in products if p["priceOnRequest"])
print(f"Parsed {len(products)} products")
print("By category:", dict(cat_counts))
print(f"Price-on-request (CALL): {call_count}")

# ── bake local images (public/products/<slug>.<ext>) ─────────────────────
import glob
IMG_DIR = os.path.join(HERE, "public", "products")
img_map = {}
if os.path.isdir(IMG_DIR):
    for f in glob.glob(os.path.join(IMG_DIR, "*")):
        base, ext = os.path.splitext(os.path.basename(f))
        if ext.lower() in (".jpg", ".jpeg", ".png", ".webp", ".avif", ".gif"):
            img_map[base] = "/products/" + os.path.basename(f)
for p in products:
    p["images"] = [img_map[p["slug"]]] if p["slug"] in img_map else []
print(f"Images matched: {sum(1 for p in products if p['images'])}/{len(products)}")

# ── emit TS ──────────────────────────────────────────────────────────────
def ts_str(s):
    return json.dumps(s if s is not None else "", ensure_ascii=False)

def ts_arr(xs):
    return "[" + ", ".join(ts_str(x) for x in xs) + "]"

ts = io.StringIO()
ts.write("// AUTO-GENERATED from FULL PRICE LIST 11062026.xlsx — do not edit by hand.\n")
ts.write("// Regenerate with: python scripts/parse-pricelist.py\n")
ts.write("import type { Product } from '@/lib/types'\n\n")
ts.write("export type CatalogProduct = Product & { priceOnRequest: boolean }\n\n")
ts.write("export const CATALOG_PRODUCTS: CatalogProduct[] = [\n")
for i, p in enumerate(products):
    tags = ", ".join(
        f"{{ id: {ts_str(f'{p['slug']}-t{j}')}, name: {ts_str(t['name'])}, "
        f"slug: {ts_str(t['slug'])}, type: '{t['type']}' }}"
        for j, t in enumerate(p["tags"]))
    specs = ", ".join(f"{ts_str(k)}: {ts_str(v)}" for k, v in p["specs"].items())
    ts.write("  {\n")
    ts.write(f"    id: {ts_str(p['slug'])}, name: {ts_str(p['name'])}, slug: {ts_str(p['slug'])},\n")
    ts.write(f"    description: {ts_str(', '.join(f'{k}: {v}' for k,v in list(p['specs'].items())[:3]) or None)},\n")
    ts.write(f"    price: {p['price']}, compare_at_price: null, priceOnRequest: {str(p['priceOnRequest']).lower()},\n")
    ts.write(f"    primary_category_id: {ts_str(p['category'])}, images: {ts_arr(p['images'])}, stock: {0 if p['priceOnRequest'] else 5},\n")
    ts.write(f"    sku: {ts_str(p['sku'])}, is_active: true, is_featured: false,\n")
    ts.write(f"    specs: {{ {specs} }}, created_at: '', updated_at: '',\n")
    ts.write(f"    tags: [{tags}],\n")
    ts.write("  },\n")
ts.write("]\n\n")
ts.write("export const CATALOG_CATEGORIES = " +
        json.dumps(sorted(cat_counts.keys()), ensure_ascii=False) + " as const\n")
os.makedirs(os.path.join(HERE, "src", "data"), exist_ok=True)
with io.open(os.path.join(HERE, "src", "data", "products.ts"), "w", encoding="utf-8") as f:
    f.write(ts.getvalue())

# ── emit SQL ─────────────────────────────────────────────────────────────
def sql_str(s):
    if s is None or s == "":
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"

brands = {}
for p in products:
    for t in p["tags"]:
        if t["type"] == "brand":
            brands[t["slug"]] = t["name"]

sql = io.StringIO()
sql.write("-- AUTO-GENERATED product seed from FULL PRICE LIST 11062026.xlsx\n")
sql.write("-- Run AFTER 001_schema.sql. Idempotent (ON CONFLICT DO NOTHING).\n\n")
sql.write("-- Brand tags\n")
vals = ",\n  ".join(f"({sql_str(n)}, {sql_str(s)}, 'brand')" for s, n in sorted(brands.items()))
sql.write(f"INSERT INTO tags (name, slug, type) VALUES\n  {vals}\nON CONFLICT (slug) DO NOTHING;\n\n")

sql.write("-- Products\n")
for p in products:
    specs_json = json.dumps(p["specs"], ensure_ascii=False).replace("'", "''")
    images_arr = ("ARRAY[" + ", ".join(sql_str(i) for i in p["images"]) + "]::text[]"
                  if p["images"] else "'{}'::text[]")
    sql.write(
        "INSERT INTO products (name, slug, description, price, primary_category_id, stock, sku, is_active, images, specs) VALUES (\n"
        f"  {sql_str(p['name'])}, {sql_str(p['slug'])}, "
        f"{sql_str(', '.join(f'{k}: {v}' for k,v in list(p['specs'].items())[:3]) or None)},\n"
        f"  {p['price']}, (SELECT id FROM categories WHERE slug = {sql_str(p['category'])}), "
        f"{0 if p['priceOnRequest'] else 5}, {sql_str(p['sku'])}, TRUE, {images_arr}, '{specs_json}'::jsonb\n"
        ") ON CONFLICT (slug) DO NOTHING;\n")
sql.write("\n-- Product ↔ brand tag links\n")
for p in products:
    bslugs = [t["slug"] for t in p["tags"] if t["type"] == "brand"]
    if not bslugs:
        continue
    arr = ", ".join(sql_str(s) for s in bslugs)
    sql.write(
        "INSERT INTO product_tags (product_id, tag_id) "
        f"SELECT p.id, t.id FROM products p JOIN tags t ON t.slug IN ({arr}) "
        f"WHERE p.slug = {sql_str(p['slug'])} ON CONFLICT DO NOTHING;\n")

os.makedirs(os.path.join(HERE, "supabase", "migrations"), exist_ok=True)
with io.open(os.path.join(HERE, "supabase", "migrations", "002_seed_products.sql"), "w", encoding="utf-8") as f:
    f.write(sql.getvalue())

print("Wrote src/data/products.ts and supabase/migrations/002_seed_products.sql")
