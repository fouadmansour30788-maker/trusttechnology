"""Parse the 2026-07 replacement catalog ("Inserted list 19-7-2026 (1).xlsx" +
"printers inserted 20-7-2026.xlsx") into products.

Emits:
  - src/data/products.ts                          (new static fallback catalog)
  - supabase/migrations/015_catalog_replacement.sql (deactivate old + insert new)

Run from the project root: python scripts/parse-new-catalog.py
"""
import re, json, io, os, glob
import openpyxl

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FILE1 = r"C:\Users\lenovo\Downloads\Inserted list 19-7-2026 (1).xlsx"
FILE2 = r"C:\Users\lenovo\Downloads\printers inserted 20-7-2026.xlsx"

# ── generic helpers ──────────────────────────────────────────────────────
def clean(v):
    if v is None:
        return None
    s = str(v).replace("™", "").replace("®", "").replace("\xa0", " ")
    s = s.replace("\n", " ; ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip(" ;,")
    return s or None

def slugify(s):
    s = re.sub(r"[^a-z0-9]+", "-", str(s).lower()).strip("-")
    return s[:80] or "item"

def parse_price(v):
    if v is None:
        return 0.0, True
    s = str(v).strip().upper()
    if "CALL" in s or s in ("", "-", "TBA", "N/A"):
        return 0.0, True
    m = re.search(r"\d+(?:\.\d+)?", str(v).replace(",", ""))
    return (float(m.group()), False) if m else (0.0, True)

def split_outside_parens(s, sep=","):
    """Split on the first `sep` that isn't nested inside parentheses —
    HP's '(Supplies: 305 Black,305 Color)' has a comma that must NOT be
    treated as the name/description boundary."""
    depth = 0
    for i, ch in enumerate(s):
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth = max(0, depth - 1)
        elif ch == sep and depth == 0:
            return s[:i].strip(), s[i + 1:].strip()
    return s, None

BRAND_KEYWORDS = [
    ("ASUS ROG", "Asus"), ("ROG STRIX", "Asus"), ("ASUS", "Asus"),
    ("LENOVO", "Lenovo"), ("THINKPAD", "Lenovo"), ("THINKBOOK", "Lenovo"),
    ("THINKCENTER", "Lenovo"), ("THINKCENTRE", "Lenovo"),
    ("DELL", "Dell"), ("HP ", "HP"), ("HP-", "HP"), ("HEWLETT", "HP"),
    ("ACER", "Acer"), ("MICROSOFT", "Microsoft"), ("SURFACE", "Microsoft"),
    ("APPLE", "Apple"), ("MSI", "MSI"), ("SEAGATE", "Seagate"),
    ("EPSON", "Epson"), ("CANON", "Canon"), ("SAMSUNG", "Samsung"),
    ("LG ", "LG"), ("OPTOMA", "Optoma"), ("APPOSTARS", "AppoStars"),
]

def detect_brand(name, default=None):
    u = name.upper()
    for kw, b in BRAND_KEYWORDS:
        if kw in u:
            return b
    return default

CPU_START_RE = re.compile(r"^(CORE|AMD|INTEL|SNAPDRAGON|APPLE M\d)", re.I)
RAM_RE = re.compile(r"^\d+\s*GB\s*DDR", re.I)
STORAGE_RE = re.compile(r"(NVME|SSD|HDD)", re.I)
SCREEN_RE = re.compile(r'(\d+(\.\d+)?\s*(-INCH|")|WUXGA|WQXGA|FHD|QHD|OLED|RETINA|PIXELSENSE|\d{3,4}\s*[xX]\s*\d{3,4})', re.I)
GPU_RE = re.compile(r"(RTX|GTX|RADEON|ARC \d|IRIS|UHD GRAPHICS|GEFORCE|CORE GPU|NEURAL ACCELERAT)", re.I)
OS_RE = re.compile(r"^(WIN\s*1[01]|DOS|MAC ?OS|MACOS)", re.I)
WARRANTY_RE = re.compile(r"^\d\s*YEARS?$", re.I)
LANG_RE = re.compile(r"^(EN|AR|BL|FR)(/(EN|AR|BL|FR))*$", re.I)

def classify_segments(segments):
    """Best-effort content classification of a laptop/desktop/Apple spec chain."""
    specs = {}
    leftovers = []
    for seg in segments:
        if not seg:
            continue
        if "CPU" not in specs and (CPU_START_RE.match(seg) or "CHIP" in seg.upper()):
            specs["CPU"] = seg
        elif "RAM" not in specs and RAM_RE.match(seg):
            specs["RAM"] = seg
        elif "RAM" not in specs and re.match(r"^\d+\s*GB\s*$", seg, re.I):
            specs["RAM"] = seg
        elif "Storage" not in specs and STORAGE_RE.search(seg):
            specs["Storage"] = seg
        elif "Language" not in specs and LANG_RE.match(seg):
            specs["Language"] = seg
        elif "Screen" not in specs and SCREEN_RE.search(seg):
            specs["Screen"] = seg
        elif "VGA" not in specs and GPU_RE.search(seg):
            specs["VGA"] = seg
        elif "Warranty" not in specs and WARRANTY_RE.match(seg):
            specs["Warranty"] = seg
        elif "OS" not in specs and OS_RE.match(seg):
            specs["OS"] = seg
        else:
            leftovers.append(seg)
    if leftovers:
        specs["Color"] = " / ".join(leftovers)
    return specs

ACRONYMS = {"CPU", "RAM", "SSD", "HDD", "OS", "DC", "LAN", "USB", "GPU"}

def label_case(label):
    words = label.split()
    return " ".join(w if w.upper() in ACRONYMS else w.title() for w in words)

def labeled_fields(text):
    """Pull 'Label: value' pairs out of free-text bullet descriptions."""
    out = {}
    for m in re.finditer(r"([A-Za-z][A-Za-z /]{2,25}):\s*([^\n;]{2,80})", text):
        label, val = clean(m.group(1)), clean(m.group(2))
        if label:
            label = label_case(label)
        if label and val and label not in out:
            out[label] = val
    return out

# ── product accumulator ──────────────────────────────────────────────────
products = []
seen_slugs = {}

def add(name, category, brand, ptype, specs, price_val, sku=None, description=None):
    name = clean(name)
    if not name or len(name) < 3:
        return
    price, on_request = parse_price(price_val)
    base_name = name if (brand and name.upper().startswith(brand.upper())) else ((brand + "-" if brand else "") + name)
    base = slugify(base_name)
    slug = base
    seen_slugs[base] = seen_slugs.get(base, 0) + 1
    if seen_slugs[base] > 1:
        slug = f"{base}-{seen_slugs[base]}"
    specs = {k: clean(v) for k, v in specs.items() if clean(v)}
    tags = []
    if brand:
        tags.append({"name": brand, "slug": slugify(brand), "type": "brand"})
    if ptype:
        tags.append({"name": ptype, "slug": slugify(ptype), "type": "type"})
    products.append({
        "name": name, "slug": slug, "category": category, "brand": brand,
        "price": price, "priceOnRequest": on_request,
        "sku": clean(sku), "specs": specs, "tags": tags,
        "description": clean(description),
    })

SPEC_FIELD_COUNT = 9  # CPU,RAM,Storage,Language,Screen,VGA,Warranty,OS,Color

def parse_blob_sheet(rows, category, ptype, sheet_brand_default=None, gaming=False):
    for row in rows:
        raw = row[0]
        price_val = row[1] if len(row) > 1 else None
        s = clean(raw)
        if not s:
            continue
        segments = [seg.strip() for seg in s.split(" , ") if seg.strip()]
        if not segments:
            continue
        first = segments[0]
        if CPU_START_RE.match(first):
            # no explicit model in this row
            model = None
            spec_segments = segments[:SPEC_FIELD_COUNT]
        else:
            model = first
            spec_segments = segments[1:1 + SPEC_FIELD_COUNT]
        specs = classify_segments(spec_segments)
        brand = detect_brand(model or s, sheet_brand_default)
        if model:
            name = model
            # Real part numbers mix letters+digits and sit at the end of the
            # model string (e.g. "...18 G835LX-S9113", "...T14 G6 21QG00APED");
            # a plain word like "EDITION" must never be picked over them.
            candidates = re.findall(r"\b([A-Z0-9]{4,}(?:[\-/][A-Z0-9]+)*)\b", model, re.I)
            skus = [c for c in candidates if re.search(r"\d", c)]
            sku = skus[-1] if skus else None
        else:
            cpu = specs.get("CPU", "")
            vga = specs.get("VGA", "")
            bits = " / ".join(x for x in [cpu, vga] if x)
            name = f"{(brand or 'Asus').upper()} GAMING LAPTOP — {bits}".strip(" —")
            sku = None
        if gaming:
            specs_ptype = "Laptop"
        else:
            specs_ptype = ptype
        add(name, category, brand, specs_ptype, specs, price_val, sku=sku)

def parse_simple_sheet(rows, category, ptype, brand_default=None, sku_prefix_split=False, long_spec_dump=False):
    for row in rows:
        raw = row[0]
        price_val = row[1] if len(row) > 1 else None
        s = clean(raw)
        if not s:
            continue
        name = s
        sku = None
        description = None
        if sku_prefix_split:
            m = re.match(r"^([A-Z0-9][A-Z0-9\.\-]{3,20})\s*,\s*(.+)$", s)
            if m:
                sku, remainder = m.group(1), m.group(2)
                if long_spec_dump:
                    # Remainder is a long bullet-point spec dump, not a short
                    # human name (e.g. AppoStars POS terminals) — keep the
                    # brand+SKU as the display name, remainder as description.
                    name = f"{brand_default} {sku}".strip() if brand_default else sku
                    description = remainder
                else:
                    name = remainder
        else:
            name, description = split_outside_parens(s)
        specs = labeled_fields(description or s)
        brand = detect_brand(name, brand_default)
        add(name, category, brand, ptype, specs, price_val, sku=sku, description=description)

# ── FILE 1: Inserted list ────────────────────────────────────────────────
wb1 = openpyxl.load_workbook(FILE1, data_only=True)
def sheet_rows(wb, name):
    ws = wb[name]
    out = []
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if vals[0] in (None, ''):
            continue
        while vals and vals[-1] is None:
            vals.pop()
        out.append(vals)
    return out

parse_blob_sheet(sheet_rows(wb1, "GAMING LAPTOPS"), "laptops", "Laptop", sheet_brand_default="Asus", gaming=True)
parse_blob_sheet(sheet_rows(wb1, "LENOVO+DELL+ASUS"), "laptops", "Laptop")
parse_blob_sheet(sheet_rows(wb1, "HP+ACER"), "laptops", "Laptop")
parse_blob_sheet(sheet_rows(wb1, "THINKPAD+THINKBOOK+SURFACE PRO"), "laptops", "Laptop", sheet_brand_default="Lenovo")
parse_blob_sheet(sheet_rows(wb1, " DESKTOPS+AIO"), "desktops", "Desktop", sheet_brand_default="Lenovo")
parse_blob_sheet(sheet_rows(wb1, "APPLE"), "laptops", "Laptop", sheet_brand_default="Apple")
parse_simple_sheet(sheet_rows(wb1, "HDD+DOCKING STATION"), "storage", "Accessory", brand_default="Seagate")
parse_simple_sheet(sheet_rows(wb1, "LENOVO ACCESSORIES"), "peripherals", "Accessory", brand_default="Lenovo", sku_prefix_split=True)

# ── FILE 2: printers inserted ────────────────────────────────────────────
wb2 = openpyxl.load_workbook(FILE2, data_only=True)
parse_simple_sheet(sheet_rows(wb2, "epson printers"), "printing", "Printer", brand_default="Epson")
parse_simple_sheet(sheet_rows(wb2, "Monitors"), "monitors", "Monitor")
parse_simple_sheet(sheet_rows(wb2, "AppoStars POS"), "pos-systems", "POS", brand_default="AppoStars", sku_prefix_split=True, long_spec_dump=True)
parse_simple_sheet(sheet_rows(wb2, "HP Printers"), "printing", "Printer", brand_default="HP")
parse_simple_sheet(sheet_rows(wb2, "Canon Printers"), "printing", "Printer", brand_default="Canon")

# ── stats ────────────────────────────────────────────────────────────────
from collections import Counter
cat_counts = Counter(p["category"] for p in products)
brand_counts = Counter(p["brand"] or "(none)" for p in products)
call_count = sum(1 for p in products if p["priceOnRequest"])
no_model_count = sum(1 for p in products if "GAMING LAPTOP —" in p["name"])
print(f"Parsed {len(products)} products")
print("By category:", dict(cat_counts))
print("By brand:", dict(brand_counts))
print(f"Price-on-request (CALL): {call_count}")
print(f"Synthesized names (no model in source): {no_model_count}")
print(f"Price range: {min(p['price'] for p in products if p['price']>0)} - {max(p['price'] for p in products)}")

# sample a few per category for spot-check
print("\n--- samples ---")
seen_cats = set()
for p in products:
    if p["category"] not in seen_cats:
        seen_cats.add(p["category"])
        print(f"[{p['category']}] {p['name']} | brand={p['brand']} | ${p['price']} | specs={p['specs']}")

# ── bake local images (none yet; placeholder pass) ───────────────────────
IMG_DIR = os.path.join(HERE, "public", "products")
img_map = {}
if os.path.isdir(IMG_DIR):
    for f in glob.glob(os.path.join(IMG_DIR, "*")):
        base, ext = os.path.splitext(os.path.basename(f))
        if ext.lower() in (".jpg", ".jpeg", ".png", ".webp", ".avif", ".gif"):
            img_map[base] = "/products/" + os.path.basename(f)
for p in products:
    p["images"] = [img_map[p["slug"]]] if p["slug"] in img_map else []
print(f"\nImages already available: {sum(1 for p in products if p['images'])}/{len(products)}")

# ── emit TS ──────────────────────────────────────────────────────────────
def ts_str(s):
    return json.dumps(s if s is not None else "", ensure_ascii=False)

def ts_arr(xs):
    return "[" + ", ".join(ts_str(x) for x in xs) + "]"

ts = io.StringIO()
ts.write("// AUTO-GENERATED from the 2026-07 replacement catalog — do not edit by hand.\n")
ts.write("// Regenerate with: python scripts/parse-new-catalog.py\n")
ts.write("import type { Product } from '@/lib/types'\n\n")
ts.write("export type CatalogProduct = Product & { priceOnRequest: boolean }\n\n")
ts.write("export const CATALOG_PRODUCTS: CatalogProduct[] = [\n")
for p in products:
    tags = ", ".join(
        f"{{ id: {ts_str(f'{p['slug']}-t{j}')}, name: {ts_str(t['name'])}, "
        f"slug: {ts_str(t['slug'])}, type: '{t['type']}' }}"
        for j, t in enumerate(p["tags"]))
    specs = ", ".join(f"{ts_str(k)}: {ts_str(v)}" for k, v in p["specs"].items())
    desc = p["description"] or (", ".join(f"{k}: {v}" for k, v in list(p["specs"].items())[:3]) or None)
    ts.write("  {\n")
    ts.write(f"    id: {ts_str(p['slug'])}, name: {ts_str(p['name'])}, slug: {ts_str(p['slug'])},\n")
    ts.write(f"    description: {ts_str(desc)},\n")
    ts.write(f"    price: {p['price']}, compare_at_price: null, priceOnRequest: {str(p['priceOnRequest']).lower()},\n")
    ts.write(f"    primary_category_id: {ts_str(p['category'])}, images: {ts_arr(p['images'])}, stock: {0 if p['priceOnRequest'] else 5},\n")
    ts.write(f"    sku: {ts_str(p['sku'])}, is_active: true, is_featured: false,\n")
    ts.write(f"    specs: {{ {specs} }}, created_at: '', updated_at: '',\n")
    ts.write(f"    tags: [{tags}],\n")
    ts.write("  },\n")
ts.write("]\n\n")
ts.write("export const CATALOG_CATEGORIES = " +
        json.dumps(sorted(cat_counts.keys()), ensure_ascii=False) + " as const\n")
os.makedirs(os.path.join(HERE, "src", "data"), exist_ok=True)
with io.open(os.path.join(HERE, "src", "data", "products.ts"), "w", encoding="utf-8") as f:
    f.write(ts.getvalue())

# ── emit SQL migration (deactivate old + insert new) ─────────────────────
def sql_str(s):
    if s is None or s == "":
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"

brands = {}
for p in products:
    for t in p["tags"]:
        if t["type"] == "brand":
            brands[t["slug"]] = t["name"]

sql = io.StringIO()
sql.write("-- AUTO-GENERATED catalog replacement from the 2026-07 price lists.\n")
sql.write("-- Old products are DEACTIVATED (is_active=false), not deleted, so order\n")
sql.write("-- history (sales_order_items etc, all ON DELETE CASCADE) is preserved.\n\n")
sql.write("UPDATE products SET is_active = false WHERE is_active = true;\n\n")

sql.write("-- Brand tags\n")
vals = ",\n  ".join(f"({sql_str(n)}, {sql_str(s)}, 'brand')" for s, n in sorted(brands.items()))
sql.write(f"INSERT INTO tags (name, slug, type) VALUES\n  {vals}\nON CONFLICT (slug) DO NOTHING;\n\n")

sql.write("-- New products\n")
for p in products:
    specs_json = json.dumps(p["specs"], ensure_ascii=False).replace("'", "''")
    desc = p["description"] or (", ".join(f"{k}: {v}" for k, v in list(p["specs"].items())[:3]) or None)
    images_arr = "'{}'::text[]"
    sql.write(
        "INSERT INTO products (name, slug, description, price, primary_category_id, stock, sku, is_active, images, specs) VALUES (\n"
        f"  {sql_str(p['name'])}, {sql_str(p['slug'])}, {sql_str(desc)},\n"
        f"  {p['price']}, (SELECT id FROM categories WHERE slug = {sql_str(p['category'])}), "
        f"{0 if p['priceOnRequest'] else 5}, {sql_str(p['sku'])}, TRUE, {images_arr}, '{specs_json}'::jsonb\n"
        ") ON CONFLICT (slug) DO UPDATE SET is_active = TRUE, price = EXCLUDED.price, stock = EXCLUDED.stock;\n")
sql.write("\n-- Product <-> brand tag links\n")
for p in products:
    bslugs = [t["slug"] for t in p["tags"] if t["type"] == "brand"]
    if not bslugs:
        continue
    arr = ", ".join(sql_str(s) for s in bslugs)
    sql.write(
        "INSERT INTO product_tags (product_id, tag_id) "
        f"SELECT p.id, t.id FROM products p JOIN tags t ON t.slug IN ({arr}) "
        f"WHERE p.slug = {sql_str(p['slug'])} ON CONFLICT DO NOTHING;\n")

os.makedirs(os.path.join(HERE, "supabase", "migrations"), exist_ok=True)
with io.open(os.path.join(HERE, "supabase", "migrations", "015_catalog_replacement.sql"), "w", encoding="utf-8") as f:
    f.write(sql.getvalue())

print("\nWrote src/data/products.ts and supabase/migrations/015_catalog_replacement.sql")
